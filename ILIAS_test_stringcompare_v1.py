# -*- coding: utf-8 -*-
"""
Created on Mon Dec 27 13:02:37 2021

@author: spauliuk

The learning platform ILIAS (https://www.ilias.de/lms-ilias-hochschulen/, https://ilias.uni-freiburg.de/ilias.php) 
offers to conduct online/remote exams via the _test_ feature.
Free text questions, where students can enter their own writing, are a crucial element of this type of exam, 
but prone to plagiarism if the exam is carried out remotely. 
This is because it is very easy to just copy-paste any text into the text fields, 
including sample answers that were circulated among students beforehand.

This script reads the xlsx export of all answers from an ILIAS test 
(for each participant, there is an own sheet with all answers)
and checks for each n(n-1)/2 pairs of students whether there are duplicate answers among ILIAS test results.

More specifically, for each question, the script creates a lower triangular matrix 
where the numerical value represents the longest common substring of the two answers given by the pair of students.
Unusually long substrings and thus overlapping answers can be easily spotted and manually checked further.
(It may turn out that students just copied part of the text of the task at hand, 
 or they actually copied their answer from another source but their own writing!)

This script is rather slow, so any improvement of its performance will be appreciated!
"""

# Import required libraries:
import openpyxl
import numpy as np
from functools import lru_cache
from operator import itemgetter

# function taken from https://www.geeksforgeeks.org/longest-common-substring-dp-29/
# on Jan 12, 2022
def longest_common_substring(x: str, y: str) -> (int, int, int):
    
    # function to find the longest common substring

    # Memorizing with maximum size of the memory as 1
    @lru_cache(maxsize=1)  
    
    # function to find the longest common prefix
    def longest_common_prefix(i: int, j: int) -> int:
      
        if 0 <= i < len(x) and 0 <= j < len(y) and x[i] == y[j]:
            return 1 + longest_common_prefix(i + 1, j + 1)
        else:
            return 0

    # diagonally computing the subproblems to decrease memory dependency
    def digonal_computation():
        
        # upper right triangle of the 2D array
        for k in range(len(x)):        
            yield from ((longest_common_prefix(i, j), i, j)
                        for i, j in zip(range(k, -1, -1), 
                                    range(len(y) - 1, -1, -1)))
        
        # lower left triangle of the 2D array
        for k in range(len(y)):        
            yield from ((longest_common_prefix(i, j), i, j)
                        for i, j in zip(range(k, -1, -1), 
                                    range(len(x) - 1, -1, -1)))

    # returning the maximum of all the subproblems
    return max(digonal_computation(), key=itemgetter(0), default=(0, 0, 0))

# open workbook with ILIAS results
NStuds = 41
'''
Below, the name of the excel files with the answer needs ot be given:
'''
mywb = openpyxl.load_workbook('NRCT_Test_Results_Jan_28_2022.xlsx')
Namesheet = mywb['Testergebnisse']
# The detailed results for each student are in a separate sheet. 
# Get the names of all these sheets first:
SheetNames = []
for m in range(2,NStuds+2):
    SheetNames.append(Namesheet.cell(m,1).value)    

'''
Extract the answers for a given question:    
This script can only evaluate one question at a time!    
'''    
#QName = 'Q12 Biofuels/agro-fuels and sustainability challenges'
#QName = 'Q14 Energy conversion vs conservation of energy'
#QName = 'Q16 Energy and sustainability connections '
#QName = 'Q20 Life cycle perspective on technology'
#QName = 'Q21 Electricity equations'
#QName = 'Q28 Extended IPAT equation'
QName = 'Hydropower 11'
#QName = 'Wind Energy SA 02'

Answers = []
for m in range(0,NStuds):
    ThisSheet = mywb[SheetNames[m]]
    n = 2
    while True:
        if ThisSheet.cell(n,2).value == QName:
            break
        n +=1
    Answers.append(ThisSheet.cell(n+1,2).value)    

# Manually fixing None type string:
if QName == 'Q21 Electricity equations':
    Answers[30] = 'No answer'
    
OverlapArray = np.zeros((NStuds,NStuds))
for m in range(0,NStuds):
    print(m)
    for n in range(0,NStuds):
        if n < m:
            if Answers[m] is not None:
                if Answers[n] is not None:
                    OverlapArray[m,n] = longest_common_substring(Answers[m],Answers[n])[0]

'''
Once done, check the OverlapArray for the longest commong substrings!
'''

#Sandbox
SheetNames[0]
[SheetNames[i] for i in[5,12,25]]


#
#
#
